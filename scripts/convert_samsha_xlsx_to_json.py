import os
import json
from openpyxl import Workbook, load_workbook

def fieldnames_from_first_row(sheet) -> list:
    fieldnames = []

    for row in sheet.rows:
        for cell in row:
            fieldnames.append(cell.value.strip())
        break
    return fieldnames


def convert_row_to_dict(row, fieldnames) -> dict:
    facilitiy = {}
    for index in range(0, len(row)):
        field_name = fieldnames[index]
        facilitiy[field_name] = row[index].value

        if facilitiy[field_name] != None:
            facilitiy[field_name] = facilitiy[field_name].strip()
    return facilitiy

def convert_sheet_to_list_of_dicts(sheet):
    fieldnames = fieldnames_from_first_row(sheet)
    skipped_header = False
    for row in sheet:
        if not(skipped_header):
            skipped_header = True
            continue
        yield convert_row_to_dict(row, fieldnames)

def apply_service_codes(facility_dict, service_code_lookup) -> dict:
    retval = {}
    category_codes = set()
    service_codes = set()
    categories_names = set()
    services_names = set()

    for key, value in facility_dict.items():
        if key.upper() in service_code_lookup:
            if facility_dict[key] == "1":
                service_code = service_code_lookup[key.upper()]

                category_codes.add(service_code['category_code'])
                categories_names.add(service_code['category_name'])

                service_codes.add(service_code['service_code'])
                services_names.add(service_code['service_name'])
        else:
            retval[key] = value

    retval['category_codes'] = list(category_codes)
    retval['service_codes'] = list(service_codes)
    retval['category_names'] = list(categories_names)
    retval['service_names'] = list(services_names)

    return retval



if __name__ == "__main__":

    WORKBOOK_FILE = os.environ['WORKBOOK_FILE']
    FACILITY_JSON_FILE = os.environ['FACILITY_JSON_FILE']

    samsha_wb = load_workbook(filename=WORKBOOK_FILE, read_only=True)

    # Fixes a bug where there isnt correct metadata on the sheet
    # https://stackoverflow.com/questions/38285364/python-openpyxl-read-only-mode-returns-a-different-row-count?rq=1
    for sheet in samsha_wb:
        sheet.max_row = sheet.max_column = None

    service_codes_sheet = samsha_wb['service code reference']

    service_code_lookup = {}
    with open('service_codes_records.json', 'w') as json_file:
        for service_code_dict in convert_sheet_to_list_of_dicts(service_codes_sheet):
            service_code_lookup[service_code_dict['service_code']] = service_code_dict
            json_file.write(json.dumps(service_code_dict))
            json_file.write('\n')


    facilities_sheet = samsha_wb['Facilities with service detail']


    # with open(FACILITY_JSON_FILE, 'w') as json_file:
    #     for facility_dict in convert_sheet_to_list_of_dicts(facilities_sheet):
    #         json_file.write(json.dumps(apply_service_codes(facility_dict, service_code_lookup)))
    #         json_file.write("\n")
